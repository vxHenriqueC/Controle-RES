import customtkinter as ctk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

ARQUIVO_EXCEL = r"C:\Users\Henrique\Downloads\eucatex\requisicoes\requisicoes.xlsx"

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

estabelecimentos = [
    "Fábrica Salto",
    "CD Logística",
    "Escritório Administrativo"
]

centros_custo = {
    "TI": "1001 - TI",
    "Financeiro": "1002 - Financeiro",
    "RH": "1003 - RH",
    "Compras": "1004 - Compras"
}

fornecedores = {
    "Dell": "0001 - Dell",
    "Microsoft": "0002 - Microsoft",
    "TOTVS": "0003 - TOTVS",
    "Oracle": "0004 - Oracle"
}

status_lista = [
    "Pendente",
    "Pago",
    "Cancelado",
    "Em Aprovação"
]

app = ctk.CTk()
app.title("Sistema Corporativo de Requisições")
app.geometry("1400x780")
app.configure(fg_color="#0f172a")

lista_requisicoes = []

# ======================================================
# CRIAR EXCEL
# ======================================================

if not os.path.exists(ARQUIVO_EXCEL):

    os.makedirs(os.path.dirname(ARQUIVO_EXCEL), exist_ok=True)

    wb = Workbook()
    ws = wb.active

    ws.title = "Requisições"

    ws.append([
        "RE",
        "Estabelecimento",
        "Centro de Custo",
        "Nota Fiscal",
        "Data Requisição",
        "Data Vencimento",
        "Valor",
        "Fornecedor",
        "Status",
        "Narrativa"
    ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:J1"

    for cell in ws[1]:

        cell.fill = PatternFill(
            start_color="2563eb",
            end_color="2563eb",
            fill_type="solid"
        )

        cell.font = Font(
            color="FFFFFF",
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    wb.save(ARQUIVO_EXCEL)

# ======================================================
# FUNÇÕES
# ======================================================

def data_atual():
    return datetime.now().strftime("%d/%m/%Y")

def limpar():

    for e in [entry_re, entry_nf, entry_vencimento, entry_valor]:
        e.delete(0, "end")

    for c in [combo_estabelecimento, combo_cc, combo_fornecedor]:
        c.set("")

    combo_status.set("Pendente")

    textbox_narrativa.delete("1.0", "end")

    entry_data_req.delete(0, "end")
    entry_data_req.insert(0, data_atual())

def atualizar_tabela():

    tabela.delete(*tabela.get_children())

    for i, item in enumerate(lista_requisicoes):

        tabela.insert(
            "",
            "end",
            iid=i,
            values=(
                item[0],
                item[1],
                item[7],
                f"R$ {item[6]:,.2f}",
                item[5],
                item[8]
            ),
            tags=(item[8],)
        )

    tabela.tag_configure("Pendente", background="#665c00")
    tabela.tag_configure("Pago", background="#14532d")
    tabela.tag_configure("Cancelado", background="#7f1d1d")
    tabela.tag_configure("Em Aprovação", background="#1e3a8a")

def incluir_re():

    re = entry_re.get().strip()

    if not re:
        return messagebox.showwarning("Aviso", "Informe a RE!")

    if any(item[0] == re for item in lista_requisicoes):
        return messagebox.showwarning("Duplicado", "Essa RE já foi adicionada!")

    try:
        valor = float(entry_valor.get().replace(",", "."))

    except:
        return messagebox.showwarning("Valor", "Informe um valor válido!")

    lista_requisicoes.append([
        entry_re.get(),
        combo_estabelecimento.get(),
        centros_custo.get(combo_cc.get(), ""),
        entry_nf.get(),
        entry_data_req.get(),
        entry_vencimento.get(),
        valor,
        str(fornecedores.get(combo_fornecedor.get(), "")),
        combo_status.get(),
        textbox_narrativa.get("1.0", "end").strip()
    ])

    atualizar_tabela()
    limpar()

def remover_re():

    item = tabela.selection()

    if not item:
        return messagebox.showwarning("Aviso", "Selecione uma RE!")

    lista_requisicoes.pop(int(item[0]))
    atualizar_tabela()

def salvar_excel():

    if not lista_requisicoes:
        return messagebox.showwarning("Aviso", "Nenhuma RE adicionada!")

    try:

        wb = load_workbook(ARQUIVO_EXCEL)
        ws = wb.active

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:J{ws.max_row}"

        cores = {
            "Pendente": "FFD966",
            "Pago": "93C47D",
            "Cancelado": "E06666",
            "Em Aprovação": "6FA8DC"
        }

        for item in lista_requisicoes:

            ws.append(item)

            linha = ws.max_row

            for cell in ws[linha]:

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            ws[f"G{linha}"].number_format = 'R$ #,##0.00'

            ws[f"I{linha}"].fill = PatternFill(
                start_color=cores[item[8]],
                end_color=cores[item[8]],
                fill_type="solid"
            )

        for coluna in ws.columns:

            tamanho = max(
                len(str(cell.value)) if cell.value else 0
                for cell in coluna
            )

            ws.column_dimensions[
                coluna[0].column_letter
            ].width = tamanho + 5

        wb.save(ARQUIVO_EXCEL)

        qtd = len(lista_requisicoes)

        lista_requisicoes.clear()

        atualizar_tabela()

        messagebox.showinfo(
            "Sucesso",
            f"{qtd} RE(s) salvas!"
        )

    except PermissionError:

        messagebox.showerror(
            "Erro",
            "Feche o Excel antes de salvar!"
        )

    except Exception as erro:

        messagebox.showerror(
            "Erro",
            str(erro)
        )

def abrir_excel():

    if os.path.exists(ARQUIVO_EXCEL):
        os.startfile(ARQUIVO_EXCEL)

# ======================================================
# TOPO
# ======================================================

topbar = ctk.CTkFrame(
    app,
    height=60,
    fg_color="#111827",
    corner_radius=0
)

topbar.pack(fill="x")

ctk.CTkLabel(
    topbar,
    text="Sistema Corporativo de Requisições",
    font=("Segoe UI", 22, "bold")
).pack(side="left", padx=25, pady=12)

# ======================================================
# CONTAINER
# ======================================================

container = ctk.CTkFrame(
    app,
    fg_color="transparent"
)

container.pack(
    fill="both",
    expand=True,
    padx=15,
    pady=15
)

frame_form = ctk.CTkScrollableFrame(
    container,
    width=420,
    fg_color="#111827",
    corner_radius=18
)

frame_form.pack(
    side="left",
    fill="y",
    padx=(0, 15)
)

frame_tabela = ctk.CTkFrame(
    container,
    fg_color="#111827",
    corner_radius=18
)

frame_tabela.pack(
    side="right",
    fill="both",
    expand=True
)

# ======================================================
# CAMPOS
# ======================================================

def label(texto):

    ctk.CTkLabel(
        frame_form,
        text=texto,
        font=("Segoe UI", 12, "bold")
    ).pack(
        anchor="w",
        padx=20,
        pady=(4, 2)
    )

def entry():

    e = ctk.CTkEntry(
        frame_form,
        width=360,
        height=34,
        corner_radius=8,
        border_width=1,
        font=("Segoe UI", 12),
        fg_color="#1e293b",
        border_color="#334155"
    )

    e.pack(
        padx=20,
        pady=(0, 5)
    )

    return e

def combo(valores):

    c = ctk.CTkComboBox(
        frame_form,
        values=valores,
        width=360,
        height=34,
        corner_radius=8,
        font=("Segoe UI", 12),
        fg_color="#1e293b",
        border_color="#334155",
        button_color="#2563eb",
        dropdown_fg_color="#1e293b"
    )

    c.pack(
        padx=20,
        pady=(0, 5)
    )

    return c

ctk.CTkLabel(
    frame_form,
    text="Nova Requisição",
    font=("Segoe UI", 20, "bold")
).pack(pady=12)

label("RE")
entry_re = entry()

label("Estabelecimento")
combo_estabelecimento = combo(estabelecimentos)

label("Centro de Custo")
combo_cc = combo(list(centros_custo.keys()))

label("Fornecedor")
combo_fornecedor = combo(list(fornecedores.keys()))

label("Nota Fiscal")
entry_nf = entry()

label("Valor")
entry_valor = entry()

label("Data Requisição")
entry_data_req = entry()
entry_data_req.insert(0, data_atual())

label("Data Vencimento")
entry_vencimento = entry()

label("Status")
combo_status = combo(status_lista)
combo_status.set("Pendente")

label("Narrativa")

textbox_narrativa = ctk.CTkTextbox(
    frame_form,
    width=360,
    height=90,
    corner_radius=8,
    font=("Segoe UI", 12),
    fg_color="#1e293b",
    border_color="#334155",
    border_width=1,
    text_color="white"
)

textbox_narrativa.pack(
    padx=20,
    pady=(0, 10)
)

# ======================================================
# BOTÕES
# ======================================================

frame_botoes = ctk.CTkFrame(
    frame_form,
    fg_color="transparent"
)

frame_botoes.pack(pady=5)

botoes = [
    ("📄 Incluir", incluir_re, "#2563eb"),
    ("❌ Remover", remover_re, "#dc2626"),
    ("💾 Salvar", salvar_excel, "#16a34a"),
    ("📂 Excel", abrir_excel, "#7c3aed")
]

for i, (txt, cmd, cor) in enumerate(botoes):

    ctk.CTkButton(
        frame_botoes,
        text=txt,
        command=cmd,
        width=165,
        height=36,
        corner_radius=10,
        fg_color=cor,
        hover_color=cor,
        font=("Segoe UI", 12, "bold")
    ).grid(
        row=i//2,
        column=i%2,
        padx=6,
        pady=6
    )

# ======================================================
# TABELA
# ======================================================

ctk.CTkLabel(
    frame_tabela,
    text="Requisições Incluídas",
    font=("Segoe UI", 22, "bold")
).pack(pady=15)

style = ttk.Style()

style.theme_use("default")

style.configure(
    "Treeview",
    background="#1e293b",
    foreground="white",
    fieldbackground="#1e293b",
    rowheight=32,
    borderwidth=0,
    font=("Segoe UI", 10)
)

style.configure(
    "Treeview.Heading",
    background="#2563eb",
    foreground="white",
    font=("Segoe UI", 11, "bold")
)

style.map(
    "Treeview",
    background=[("selected", "#3b82f6")]
)

colunas = (
    "RE",
    "Estabelecimento",
    "Fornecedor",
    "Valor",
    "Vencimento",
    "Status"
)

tabela = ttk.Treeview(
    frame_tabela,
    columns=colunas,
    show="headings",
    height=18
)

for c in colunas:

    tabela.heading(c, text=c)

    tabela.column(
        c,
        width=150,
        anchor="center"
    )

tabela.pack(
    fill="both",
    expand=True,
    padx=15,
    pady=(0, 15)
)

app.mainloop()